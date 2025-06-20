"""app/ui/pages/loader_page.py – Araç Yükleme (QR'li)
=====================================================
• Barkod okutuldukça paket `loaded=1`, `loaded_by`, `loaded_time` güncellenir.
• Liste yalnızca **en az bir paketi yüklenmiş** sevkiyatları gösterir.
• "Liste Yazdır (QR)" butonu: sevkiyat başlığına `qr_token` üretir, QR kodlu PDF oluşturur.
"""
from __future__ import annotations
from textwrap import wrap
from reportlab.lib.pagesizes import landscape, A4
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from textwrap import wrap
from reportlab.pdfbase.pdfmetrics import stringWidth
import app.settings as st

import csv, os, io, uuid, getpass
from pathlib import Path
from datetime import date
from typing import Dict, List
from reportlab.lib.utils import ImageReader 
from PyQt5.QtCore    import Qt, QDate
from PyQt5.QtGui     import QCursor
from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QDateEdit, QLineEdit,
    QPushButton, QTableWidget, QTableWidgetItem, QHeaderView, QMessageBox,
    QFileDialog, QMenu,QDialog, QListWidget, QListWidgetItem, QAbstractItemView
)
from PyQt5.QtCore import Qt, QDate, QTimer      # ← QTimer eklendi


BASE_DIR = Path(__file__).resolve().parents[3]
import sys; sys.path.append(str(BASE_DIR))

from app.shipment import (
    list_headers_range, trip_by_barkod,
    mark_loaded, set_trip_closed
)
from app import toast
from app.dao.logo import exec_sql, ensure_qr_token, fetch_all, fetch_one

import qrcode

from PyQt5.QtMultimedia import QSoundEffect
from PyQt5.QtCore import QUrl
SOUND_DIR = BASE_DIR / "sounds"

def _load_wav(name):
    s = QSoundEffect()
    s.setSource(QUrl.fromLocalFile(str(SOUND_DIR / name)))
    s.setVolume(0.9)
    return s

snd_ok   = _load_wav("ding.wav")
snd_dupe = _load_wav("bip.wav")
snd_err  = _load_wav("error.wav")

from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm

# ───────────────────────── Tablo kolonları
COLS = [
    ("id",           "#"),
    ("order_no",     "Sipariş"),
    ("customer_code","Cari Kod"),
    ("customer_name","Müşteri"),
    ("region",       "Bölge"),
    ("address1",     "Adres"),
    ("pkgs_total",   "Paket"),
    ("pkgs_loaded",  "Yüklendi"),
    ("loaded_at",    "Yüklendi 🕒"),
    ("status_txt",   "Durum"),
]


# >>>>> EKLE >>>>>
class ColumnSelectDialog(QDialog):
    """Excel/CSV’de hangi kolonlar olsun?"""
    def __init__(self, parent, cols):
        super().__init__(parent)
        self.setWindowTitle("Kolon Seç")
        self.resize(250, 300)
        v = QVBoxLayout(self)

        self.lst = QListWidget(selectionMode=QAbstractItemView.MultiSelection)
        for key, header in cols:
            itm = QListWidgetItem(header)
            itm.setData(Qt.UserRole, key)
            itm.setSelected(True)           # varsayılan: hepsi
            self.lst.addItem(itm)
        v.addWidget(self.lst)

        btn_ok = QPushButton("Tamam"); btn_ok.clicked.connect(self.accept)
        v.addWidget(btn_ok, alignment=Qt.AlignRight)

    def selected_keys(self):
        return [i.data(Qt.UserRole) for i in self.lst.selectedItems()]

def _ask_columns(parent) -> list[str] | None:
    dlg = ColumnSelectDialog(parent, COLS)
    return dlg.selected_keys() if dlg.exec_() else None
# <<<<< EKLE <<<<<

# ════════════════════════ UI ═══════════════════════════════════
class LoaderPage(QWidget):
    def __init__(self):
        super().__init__()
        self._build_ui()
      # ► Otomatik yenileme – her 30 sn
        self._timer = QTimer(self)
        self._timer.timeout.connect(self.refresh)
        self._timer.start(st.get("loader.auto_refresh", 30) * 1000)         # 30 000 ms = 30 sn

    def showEvent(self, event):              # <– EKLENDİ
        """Sekmeye/ekrana dönüldüğünde barkod girişine odaklan."""
        super().showEvent(event)
        QTimer.singleShot(0, self.entry.setFocus)

    def _build_ui(self):
        lay = QVBoxLayout(self)
        lay.addWidget(QLabel("<b>Araç Yükleme</b>"))

        # — filtre barı —
        top = QHBoxLayout()
        self.dt_from = QDateEdit(QDate.currentDate()); top.addWidget(QLabel("Baş:")); top.addWidget(self.dt_from)
        self.dt_to   = QDateEdit(QDate.currentDate()); top.addWidget(QLabel("Bitiş:")); top.addWidget(self.dt_to)
        self.search  = QLineEdit(); self.search.setPlaceholderText("Ara… (sipariş/cari/bölge)")
        top.addWidget(self.search, 1)
        btn_list   = QPushButton("Yüklemeleri Getir"); btn_list.clicked.connect(self.refresh)
        btn_csv    = QPushButton("Excel/CSV");        btn_csv.clicked.connect(self.export_csv)
        btn_print  = QPushButton("Liste Yazdır (QR)"); btn_print.clicked.connect(self.print_loading_list)  # ★
        btn_done   = QPushButton("Yükleme Tamam");    btn_done.clicked.connect(self.close_trip)
        top.addStretch(); top.addWidget(btn_list); top.addWidget(btn_csv); top.addWidget(btn_print); top.addWidget(btn_done)
        lay.addLayout(top)

        # — tablo —
        self.tbl = QTableWidget(0, len(COLS))
        self.tbl.setHorizontalHeaderLabels([h for _k, h in COLS])
        self.tbl.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.tbl.setSortingEnabled(True)
        self.tbl.setSelectionBehavior(QTableWidget.SelectRows)
        self.tbl.setContextMenuPolicy(Qt.CustomContextMenu)
        self.tbl.customContextMenuRequested.connect(self._ctx_menu)
        lay.addWidget(self.tbl)

        # — barkod entry —
        bar = QHBoxLayout()
        self.entry = QLineEdit(); self.entry.setPlaceholderText("Paket barkodu → Enter")
        self.entry.returnPressed.connect(self.on_scan); bar.addWidget(self.entry)
        lay.addLayout(bar)


    def showEvent(self, event):
        super().showEvent(event)

        # ► her gösterimde barkod kutusuna odak
        self.entry.setFocus(Qt.OtherFocusReason)

        # ► otomatik yenileme yeniden başlasın
        self._timer.start()

  
    # ══════════════ Veri yükle & tablo doldur ═══════════════════
    def refresh(self):
        d1 = self.dt_from.date().toPyDate().isoformat()
        d2 = self.dt_to.date().toPyDate().isoformat()
        rows = list_headers_range(d1, d2)

        # Arama filtresi
        q = self.search.text().strip().upper()
        if q:
            rows = [r for r in rows if q in r["order_no"].upper()
                               or q in (r["customer_code"] or "").upper()
                               or q in (r["region"] or "").upper()]

        # Yalnızca en az 1 paket yüklenmişse göster
        rows = [r for r in rows if r["pkgs_loaded"] > 0]

        # Başlık satırı ikon-metni
        for r in rows:
            r["status_txt"] = (
                "🚚" if r.get("en_route")                      # araç yolda
                else "✔" if r["closed"]                       # tamamen yüklü
                else "⏳"                                      # bekliyor
            )
            r["loaded_at"] = (r.get("loaded_at") or "")[:19]

        # Tabloyu güncelle
        self._rows   = rows
        self._id_map = {r["id"]: r for r in rows}             # 🔸 fix: tüm id’ler
        self.tbl.setRowCount(0)
        for rec in rows:
            self._add_row(rec)  
        self.entry.setFocus(Qt.OtherFocusReason)
    # ────────────────────────────────────────────────────────────


    def _add_row(self, rec: Dict):
        r = self.tbl.rowCount(); self.tbl.insertRow(r)
        for c, (k, _h) in enumerate(COLS):
            itm = QTableWidgetItem(str(rec.get(k, "")))
            itm.setTextAlignment(Qt.AlignCenter)
            # renk mantığı
            if rec["pkgs_loaded"] >= rec["pkgs_total"]:
                itm.setBackground(Qt.green)
            elif rec["pkgs_loaded"] == 0:
                itm.setBackground(Qt.red)
            self.tbl.setItem(r, c, itm)



    # ══════════════ Barkod okuma ════════════════════════════════
    def on_scan(self):
        raw = self.entry.text().strip()
        self.entry.clear()
        if not raw or "-K" not in raw:
            snd_err.play()                                  # 🔊 hata
            return

        inv_root, pkg_txt = raw.rsplit("-K", 1)
        try:
            pkg_no = int(pkg_txt)
        except ValueError:
            snd_err.play()                                  # 🔊 hata
            return

        # ► Aktif sevkiyat başlığını bul
        trip = trip_by_barkod(inv_root)          # tarih filtresiz
        if not trip:
            snd_err.play()                                  # 🔊 hata
            QMessageBox.warning(self, "Paket", "Sevkiyat başlığı bulunamadı!")
            return

        trip_id, pkg_tot = trip
        if not (1 <= pkg_no <= pkg_tot):
            snd_err.play()                                  # 🔊 hata
            QMessageBox.warning(self, "Paket", "Paket numarası toplamı aşmış!")
            return

        # ──────────────────────────────────────────────
        # 1) shipment_loaded + shipment_header güncelle
        # ──────────────────────────────────────────────
        ok = mark_loaded(trip_id, pkg_no)
        if ok == 0:                 # yinelenen okuma
            snd_dupe.play()                             # 🔊 tekrar
            toast("Uyarı", "Bu paket zaten yüklenmiş!")
            return

        # shipment_lines güncellemesi artık mark_loaded() içinde yapılıyor

        snd_ok.play()                                   # 🔊 başarılı okuma
        toast("Paket Yüklendi", f"{inv_root} K{pkg_no}")
        self.refresh()

    # ════════════════════════════════════════════════════════════
    # ───────────── Uygulama Ayarları Anında Uygula ─────────────
    def apply_settings(self):
        """MainWindow -> _apply_global_settings çağırır."""
        # ► Otomatik yenile
        self._timer.setInterval(st.get("loader.auto_refresh", 30) * 1000)

        # ► Otomatik fokus
        self._auto_focus = st.get("ui.auto_focus", True)

        # ► Ses
        vol = st.get("ui.sounds.volume", 0.9)
        enabled = st.get("ui.sounds.enabled", True)
        for s in (snd_ok, snd_dupe, snd_err):
            s.setVolume(vol if enabled else 0.0)



       # ════════════════════════════════════════════════════════════
    # ════════════════════════════════════════════════════════════
    def print_loading_list(self):
        """
        Seçili satırlar PDF’e basılır; seçim yoksa ekrandaki tüm satırlar basılır.
        Tabloyu hangi kolona göre sıraladıysanız PDF de o sırayla oluşturulur.
        Sayfa dolduğunda otomatik yeni sayfa açılır.
        """
        # -------------------------------------------------------- #
        # 0) Tablo verisi var mı?                                  #
        # -------------------------------------------------------- #
        if not getattr(self, "_rows", None):
            QMessageBox.warning(self, "Liste", "Önce listeyi getir!")
            return

        # -------------------------------------------------------- #
        # 1) Ekrandaki görünür sırayı çıkart                       #
        # -------------------------------------------------------- #
        visible_ids = [
            int(self.tbl.item(r, 0).text())  for r in range(self.tbl.rowCount())
        ]
        rows_in_view = [self._id_map[i] for i in visible_ids]

        # -------------------------------------------------------- #
        # 2) Satır seçimi varsa filtre uygula                      #
        # -------------------------------------------------------- #
        sel_rows = {ix.row() for ix in self.tbl.selectionModel().selectedRows()}
        rows_to_print = [rows_in_view[r] for r in sel_rows] if sel_rows else rows_in_view
        if not rows_to_print:
            QMessageBox.information(self, "Liste", "Basılacak satır yok.")
            return

        # -------------------------------------------------------- #
        # 3) Çıkış dosya adı (tarih-saat)                          #
        # -------------------------------------------------------- #
        out_pdf = BASE_DIR / "output" / f"loader_{date.today():%Y%m%d_%H%M%S}.pdf"

        # -------------------------------------------------------- #
        # 4) Font & PDF ayarı                                      #
        # -------------------------------------------------------- #
        try:
            font_path = BASE_DIR / "fonts" / "DejaVuSans.ttf"
            pdfmetrics.registerFont(TTFont("DejaVuSans", str(font_path)))
            FONT = "DejaVuSans"
        except Exception:
            FONT = "Helvetica"

        W, H = landscape(A4)
        pdf   = canvas.Canvas(str(out_pdf), pagesize=(W, H))
        pdf.setFont(FONT, 8)

        cols = [
            ("QR",        22*mm), ("Sipariş",   28*mm),
            ("Cari Kod",  24*mm), ("Müşteri",   38*mm),
            ("Bölge",     28*mm), ("Adres",     50*mm),
            ("Paket",     10*mm), ("Yüklendi",  32*mm),
            ("Kaşe",      40*mm),
        ]
        margin, header_h, row_h_min = 15*mm, 12*mm, 24*mm
        y_top = H - margin

        # 🔸 Toplam koli hesapla
        total_pkgs = sum(r["pkgs_total"] for r in rows_to_print)

        # Yardımcı: kelime sar
        def split_text(txt, font, size, max_w):
            out, cur = [], ""
            for w in str(txt).split():
                test = (cur + " " + w).strip()
                if stringWidth(test, font, size) <= max_w:
                    cur = test
                else:
                    if cur: out.append(cur); cur = w
            out.append(cur); return out

        def draw_header(y):
            # 🔸 Üst başlık satırı
            pdf.setFont(FONT, 10)
            pdf.drawString(margin,
                           y + 4*mm,
                           f"Tarih: {date.today():%d.%m.%Y}    "
                           f"Toplam Koli: {total_pkgs}")
            pdf.setFont(FONT, 8)
            x = margin
            for title, w in cols:
                pdf.rect(x, y-header_h, w, header_h)
                pdf.drawCentredString(x + w/2, y-header_h + 3, title)
                x += w

        # -------------------------------------------------------- #
        # 5) Başlık + satırlar                                    #
        # -------------------------------------------------------- #
        draw_header(y_top); y_cursor = y_top - header_h

        for rec in rows_to_print:
            # ― QR ―
            buf = io.BytesIO()
            qrcode.make(ensure_qr_token(rec["order_no"])).save(buf, "PNG")
            qr_img = ImageReader(buf); buf.seek(0)

            cell_vals = [
                rec["order_no"], rec["customer_code"], rec["customer_name"],
                rec["region"], rec["address1"],
                f"{rec['pkgs_loaded']} / {rec['pkgs_total']}",
                rec["loaded_at"][:19], "",
            ]

            dyn_row_h, cell_lines = row_h_min, []
            for (_t, w), txt in zip(cols[1:], cell_vals):
                lines = split_text(txt, FONT, 7, w-4*mm)
                cell_lines.append(lines)
                dyn_row_h = max(dyn_row_h, 6 + 9*len(lines))

            if y_cursor - dyn_row_h < margin:
                pdf.showPage(); pdf.setFont(FONT, 8)
                draw_header(H - margin)
                y_cursor = H - margin - header_h

            # çerçeveler
            x = margin
            for _t, w in cols:
                pdf.rect(x, y_cursor-dyn_row_h, w, dyn_row_h); x += w

            # QR
            qr_sz = 18*mm
            pdf.drawImage(
                qr_img,
                margin + (cols[0][1]-qr_sz)/2,
                y_cursor - dyn_row_h + (dyn_row_h-qr_sz)/2,
                qr_sz, qr_sz, preserveAspectRatio=True
            )

            # Metin
            x = margin + cols[0][1]; pdf.setFont(FONT, 7)
            for (_t, w), lines in zip(cols[1:], cell_lines):
                for i, line in enumerate(lines):
                    pdf.drawString(x+2, y_cursor - 9 - i*9, line)
                x += w

            y_cursor -= dyn_row_h

        # -------------------------------------------------------- #
        pdf.save()
        os.startfile(out_pdf)            # Windows: PDF görüntüleyici
        toast("PDF Hazır", str(out_pdf))
        





    def split_text(text: str, font_name: str, font_size: int, max_width: float):
        """
        max_width (pt) değerini aşmadan kelimeleri satırlara ayır.
        """
        words  = text.split()
        lines  = []
        line   = ""
        for w in words:
            test = (line + " " + w).strip()
            if stringWidth(test, font_name, font_size) <= max_width:
                line = test
            else:
                if line:        # önceki satırı kaydet
                    lines.append(line)
                line = w        # kelimeyi yeni satıra taşı
        lines.append(line)
        return lines
    # ══════════════ Manuel kapama ═══════════════════════════════
    def close_trip(self):
        """
        Seçili sevkiyat(lar)ı kapatır.
        • Eksik koli varsa önce onay ister.
        • Eksik kapatma USER_ACTIVITY tablosuna loglanır.
        """
        rows = {i.row() for i in self.tbl.selectedIndexes()}
        if not rows:
            return

        for row in rows:
            trip_id = int(self.tbl.item(row, 0).text())
            rec = self._id_map.get(trip_id)
            if not rec:
                continue

            # Eksik koli var mı?
            if rec["pkgs_loaded"] < rec["pkgs_total"]:
                ans = QMessageBox.question(
                    self, "Eksik Koli",
                    f"{rec['pkgs_loaded']} / {rec['pkgs_total']} yüklendi.\n"
                    "Yine de 'Yükleme Tamam' yapılsın mı?",
                    QMessageBox.Yes | QMessageBox.No
                )
                if ans == QMessageBox.No:
                    continue  # kullanıcı vazgeçti

                # Log – eksik kapatma
                exec_sql("""
                    INSERT INTO USER_ACTIVITY
                        (username, action, details, order_no)
                    SELECT ?, 'TRIP_MANUAL_CLOSED_INCOMPLETE', ?, order_no
                      FROM shipment_header
                     WHERE id = ?""",
                    getpass.getuser(),
                    f"{rec['pkgs_loaded']}/{rec['pkgs_total']}",
                    trip_id
                )

            # Kapama işlemi (en_route = 1)
            set_trip_closed(trip_id, True)

        self.refresh()


   # ─────────────────────────────────────────────────────────────
#  Dışa Aktarım  –  CSV / Excel  (kolon seçmeli)
# ─────────────────────────────────────────────────────────────
    def export_csv(self):
        """Mevcut satırları CSV / Excel’e dışa aktarır.
        • Önce kolon seçimi diyalogu açılır.
        • Seçim yapılmazsa (İptal) işlem durur.
        """
        if not getattr(self, "_rows", None):
            QMessageBox.warning(self, "Dışa Aktarım", "Önce listeyi getir!"); return

        sel_keys = _ask_columns(self)                 # ← yeni diyalog
        if not sel_keys:                              # İptal
            return

        fn, _ = QFileDialog.getSaveFileName(
            self, "Kaydet", str(BASE_DIR / f"loader_{date.today():%Y%m%d}"),
            "Excel (*.xlsx);;CSV (*.csv)"
        )
        if not fn:
            return

        if fn.lower().endswith(".csv"):
            self._write_csv(fn, sel_keys)
        else:
            self._write_xlsx(fn, sel_keys)

        QMessageBox.information(self, "Dışa Aktarım", f"Dosya yazıldı:\n{fn}")

        # ---------------- CSV -------------------------------------
    def _write_csv(self, path: str, keys: list[str]):
        """
        Seçili kolonları (‘keys’) kullanarak CSV oluşturur ve
        tamamlandığında varsayılan programla dosyayı açar.
        """
        with open(path, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow([header for k, header in COLS if k in keys])   # başlık
            for rec in self._rows:
                w.writerow([rec.get(k, "") for k, _h in COLS if k in keys])

        os.startfile(path)   # ↻  otomatik aç

    # ---------------- XLSX ------------------------------------
    def _write_xlsx(self, path: str, keys: list[str]):
        """
        Seçili kolonlarla Excel (.xlsx) üretir; bittiğinde otomatik açar.
        """
        try:
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter
        except ImportError:
            QMessageBox.warning(self, "Excel", "pip install openpyxl")
            return

        wb = Workbook(); ws = wb.active

        ws.append([header for k, header in COLS if k in keys])        # başlık
        for rec in self._rows:                                        # satırlar
            ws.append([rec.get(k, "") for k, _h in COLS if k in keys])

        # Otomatik sütun genişliği
        for col_idx in range(1, len(keys) + 1):
            max_len = max(
                len(str(ws.cell(row=r, column=col_idx).value) or "")
                for r in range(1, ws.max_row + 1)
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2

        wb.save(path)
        os.startfile(path)   # ↻  otomatik aç


        

    # ══════════════ Sağ‑tık Detay ═══════════════════════════════
    def _ctx_menu(self, pos):
        idx = self.tbl.indexAt(pos); row = idx.row()
        if row < 0:
            return
        trip_id = int(self.tbl.item(row, 0).text())
        rec = self._id_map.get(trip_id)
        if not rec:
            return
        txt = [f"<b>Sipariş No</b>: {rec['order_no']}"]
        for k in ("customer_code", "customer_name", "region", "address1",
                  "pkgs_total", "pkgs_loaded", "loaded_at", "closed", "created_at"):
            txt.append(f"{k.replace('_',' ').title()}: {rec.get(k, '')}")
        QMessageBox.information(self, "Sipariş Detay", "<br>".join(txt))
